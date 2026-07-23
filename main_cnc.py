import sys
import serial  # Untuk komunikasi lewat kabel USB ke Arduino
import time  # Untuk mengatur jeda waktu (delay)
import re  # Untuk mencari angka di dalam teks G-Code secara cerdas
import os  # Untuk mengatur file di dalam sistem komputer (seperti menghapus gambar sementara)
import datetime  # Untuk mengambil data jam dan tanggal hari ini
import queue  # Untuk mengirim event sensor dengan aman antar-thread
from pathlib import Path
import openpyxl  # Untuk membuat dan mengedit file Excel
from openpyxl.drawing.image import Image as ExcelImage  # Untuk memasukkan gambar ke dalam Excel

import matplotlib  # Pustaka utama untuk menggambar grafik

matplotlib.use('qtagg')  # Mengatur agar grafik bisa digabungkan dengan aplikasi PyQt6

# Mengimpor komponen-komponen untuk membuat antarmuka aplikasi (tombol, teks, kotak, dll)
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QProgressBar, QMessageBox, QFileDialog, QDialog, QFrame,
                             QSizePolicy, QLineEdit, QGridLayout)
from PyQt6.QtCore import QThread, pyqtSignal, Qt
from PyQt6.QtGui import QFont, QPixmap, QColor

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# ==========================================
# KONFIGURASI PORT KABEL USB
# Pastikan nomor COM di bawah ini sama dengan yang ada di Device Manager laptop Anda
# ==========================================
PORT_GRBL = 'COM12'  # Colokan USB yang menuju ke Arduino penggerak mesin (GRBL)
PORT_SENSOR = 'COM9'  # Colokan USB yang menuju ke Arduino pembaca sensor (AR3)
BAUD_RATE = 115200  # Kecepatan transfer data (harus sama dengan di program Arduino)
DEFAULT_GCODE_FILE = Path(__file__).with_name("KotakSaja-B_Cu.gbr_iso_combined_cnc_flatcam.txt")


# ==========================================
# 0. PEKERJA LATAR BELAKANG: PROSES KALIBRASI
# Menggunakan QThread agar saat kalibrasi berjalan, aplikasi tidak macet/hang
# ==========================================
class CalibrationSensorReader(QThread):
    """Membaca port sensor selama kalibrasi tanpa berebut dengan worker GRBL."""

    def __init__(self, sensor):
        super().__init__()
        self.sensor = sensor
        self.is_running = True
        self.events = queue.Queue()
        self.bracket_pernah_terpasang = False
        self.bracket_terpasang = False
        self.casing_tertutup = False

    def proses_data(self, data):
        if data == "LSBON":
            self.bracket_pernah_terpasang = True
            self.bracket_terpasang = True
        elif data == "LSBOFF":
            self.bracket_terpasang = False
        elif data == "LSC_ON":
            self.casing_tertutup = True
        elif data == "LSC_OFF":
            self.casing_tertutup = False
        self.events.put(data)

    def ambil_event(self, event_yang_dicari=None):
        while True:
            try:
                data = self.events.get_nowait()
            except queue.Empty:
                return None
            if event_yang_dicari is None or data in event_yang_dicari:
                return data

    def bersihkan_event(self):
        while self.ambil_event() is not None:
            pass

    def run(self):
        while self.is_running:
            try:
                if self.sensor.in_waiting:
                    raw = self.sensor.readline().decode(errors="ignore").strip()
                    if raw.startswith("#") and raw.endswith(";"):
                        self.proses_data(raw[1:-1])
                    self.sensor.write(b"ACK\n")
                else:
                    time.sleep(0.01)
            except Exception as error:
                print(f"Pembaca sensor kalibrasi berhenti: {error}")
                return

    def stop(self):
        self.is_running = False


class CalibrationWorker(QThread):
    # Sinyal-sinyal untuk lapor ke layar utama (update tulisan, nilai loading, dll)
    progress_update = pyqtSignal(int)
    status_update = pyqtSignal(str)
    finished_success = pyqtSignal(object, object)
    finished_mock = pyqtSignal()
    failed = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.grbl = None
        self.sensor = None
        self.sensor_reader = None
        self.is_running = True

    # Fungsi khusus untuk mengirim perintah G-Code dasar ke mesin
    def kirim(self, cmd):
        print(f"> {cmd}")
        if self.grbl:
            self.grbl.write((cmd + '\n').encode())  # Ubah teks jadi kode mesin (bytes)
            self.grbl.flush()

    # Menunggu mesin menjawab "ok", artinya mesin sudah selesai bergerak
    def wait_grbl_ok(self, timeout_seconds=10):
        deadline = time.monotonic() + timeout_seconds
        while self.is_running and time.monotonic() < deadline:
            if not self.bracket_aman():
                return False
            try:
                response = self.grbl.readline().decode(errors="ignore").strip()
                if response.lower() == "ok":
                    return True
                if "error" in response.lower() or "alarm" in response.lower():
                    print(f"GRBL gagal: {response}")
                    return False
            except Exception as error:
                print(f"Gagal membaca GRBL: {error}")
                return False
            time.sleep(0.01)
        return False

    # Fungsi untuk membaca kiriman data rahasia (#...;) dari Arduino sensor
    def baca_sensor(self, event_yang_dicari=None):
        if self.sensor_reader:
            return self.sensor_reader.ambil_event(event_yang_dicari)
        while self.sensor.in_waiting:
            raw = self.sensor.readline().decode(errors="ignore").strip()
            # Jika teks diawali '#' dan diakhiri ';', buang simbol tersebut dan ambil isinya
            if raw.startswith("#") and raw.endswith(";"):
                data = raw[1:-1]
                if event_yang_dicari is None or data in event_yang_dicari:
                    return data
        return None

    def bracket_aman(self):
        if (self.sensor_reader and self.sensor_reader.bracket_pernah_terpasang
                and not self.sensor_reader.bracket_terpasang):
            try:
                self.grbl.write(b"!")
                time.sleep(0.1)
                self.grbl.write(b"\x18")
            except Exception as error:
                print(f"Gagal menghentikan GRBL saat bracket lepas: {error}")
            self.is_running = False
            self.gagal("Bracket PCB terlepas. Kalibrasi dihentikan.")
            return False
        return True

    def stop_sensor_reader(self):
        if self.sensor_reader:
            self.sensor_reader.stop()
            self.sensor_reader.wait(500)
            self.sensor_reader = None

    def gagal(self, pesan):
        print(f"Kalibrasi gagal: {pesan}")
        self.status_update.emit(pesan)
        self.failed.emit(pesan)
        self.stop_sensor_reader()

    # Misi utama yang dikerjakan oleh pekerja kalibrasi ini
    def run(self):
        self.status_update.emit("Mencari Mesin di Port Serial...")
        self.progress_update.emit(5)  # Loading bar naik 5%
        time.sleep(0.5)

        mock_mode = False
        # Coba hubungkan ke port USB
        try:
            self.grbl = serial.Serial(PORT_GRBL, BAUD_RATE, timeout=1)
            self.sensor = serial.Serial(PORT_SENSOR, BAUD_RATE, timeout=1)
        except Exception as e:
            # Jika gagal (kabel belum dicolok), masuk ke mode Simulasi
            print(f"[AUTO-SIMULASI] Mesin tidak ditemukan: {e}")
            mock_mode = True

        # JIKA MESIN ASLI TERHUBUNG:
        if not mock_mode:
            try:
                self.sensor.reset_input_buffer()
                self.sensor_reader = CalibrationSensorReader(self.sensor)
                self.sensor_reader.start()

                # 1. Pastikan PCB sudah dijepit
                self.status_update.emit("Menunggu PCB Terpasang...")
                self.progress_update.emit(10)

                lsb_ok = False
                while self.is_running and not lsb_ok:
                    data = self.baca_sensor({"LSBON"})
                    if data == "LSBON":  # Kode LSBON berarti limit switch bracket tertekan
                        lsb_ok = True
                    time.sleep(0.1)

                if not self.is_running: return  # Berhenti jika user klik silang (X)

                # 2. Bangunkan mesin GRBL
                self.status_update.emit("Inisialisasi GRBL...")
                self.grbl.write(b"\r\n\r\n")
                time.sleep(2)
                self.grbl.flushInput()
                self.sensor_reader.bersihkan_event()

                # Buka kunci alarm ($X), pakai satuan mm (G21), pakai sistem maju pelan (G91)
                self.kirim("$X");
                self.kirim("G21");
                self.kirim("G91")

                # 3. Cari titik batas ujung mesin (Limit Switch Hardware)
                self.status_update.emit("Mencari Limit Switch X, Y, Z...")
                self.progress_update.emit(30)
                limx, limy, limz = False, False, False

                while self.is_running:
                    if not self.bracket_aman():
                        return
                    # Buat perintah jalan pelan-pelan
                    move_cmd = "G1 "
                    if not limx: move_cmd += "X-2 "  # X gerak ke kiri
                    if not limy: move_cmd += "Y-2 "  # Y gerak ke bawah
                    if not limz: move_cmd += "Z2 "  # Z gerak ke atas (mengamankan pahat)
                    move_cmd += "F300"  # Kecepatan lambat

                    # Kirim perintah geraknya
                    if move_cmd != "G1 F300":
                        self.kirim(move_cmd)
                    time.sleep(0.05)

                    # Dengar teriakan sensor, adakah limit switch yang mentok?
                    data = self.baca_sensor({"LIMX", "LIMY", "LIMZ"})
                    if data == "LIMX":
                        limx = True
                        self.grbl.write(b'\x18');
                        time.sleep(1)  # Soft Reset agar berhenti mendadak
                        self.kirim("$X");
                        self.kirim("G21");
                        self.kirim("G91")
                    elif data == "LIMY":
                        limy = True
                        self.grbl.write(b'\x18');
                        time.sleep(1)
                        self.kirim("$X");
                        self.kirim("G21");
                        self.kirim("G91")
                    elif data == "LIMZ":
                        limz = True
                        self.grbl.write(b'\x18');
                        time.sleep(1)
                        self.kirim("$X");
                        self.kirim("G21");
                        self.kirim("G91")

                    # Jika ketiga limit switch sudah tersentuh, keluar dari pencarian
                    if limx and limy and limz:
                        self.grbl.write(b'\x18');
                        time.sleep(1)
                        self.kirim("$X");
                        self.kirim("G21");
                        self.kirim("G91")
                        self.sensor_reader.bersihkan_event()
                        time.sleep(0.2)
                        break

                if not self.is_running: return
                if not self.bracket_aman(): return

                # 4. Berpindah ke Titik Nol PCB yang Sebenarnya (Sumbu X)
                self.status_update.emit("Geser X 37mm...")
                self.progress_update.emit(50)
                self.kirim("G91");
                self.kirim("G1 X37 F300")  # Geser X sejauh 37mm
                if not self.wait_grbl_ok():
                    if self.is_running:
                        self.gagal("GRBL tidak mengonfirmasi pergeseran X.")
                    return
                self.kirim("G92 X0");
                if not self.wait_grbl_ok():
                    if self.is_running:
                        self.gagal("GRBL tidak mengonfirmasi titik nol X.")
                    return

                # 5. Berpindah ke Titik Nol PCB yang Sebenarnya (Sumbu Y)
                self.status_update.emit("Geser Y 118mm...")
                self.progress_update.emit(70)
                self.kirim("G91");
                self.kirim("G1 Y118 F300")  # Geser Y sejauh 118mm
                if not self.wait_grbl_ok():
                    if self.is_running:
                        self.gagal("GRBL tidak mengonfirmasi pergeseran Y.")
                    return
                self.kirim("G92 Y0");
                if not self.wait_grbl_ok():
                    if self.is_running:
                        self.gagal("GRBL tidak mengonfirmasi titik nol Y.")
                    return

                # 6. Turunkan pahat berlahan sampai menyentuh PCB (Sumbu Z)
                self.status_update.emit("Menyesuaikan Home Z (PCB Sentuh)...")
                self.progress_update.emit(85)
                self.sensor_reader.bersihkan_event()
                pcb_tersentuh = False
                while self.is_running:
                    if not self.bracket_aman():
                        return
                    self.kirim("G1 Z-2 F300")  # Pahat turun pelan-pelan
                    data = self.baca_sensor({"PCBON"})
                    if data == "PCBON":  # Sensor mendeteksi ada arus mengalir dari pahat ke PCB
                        pcb_tersentuh = True
                        self.grbl.write(b'\x18');
                        time.sleep(1)  # Berhenti mendadak!
                        self.kirim("$X");
                        self.kirim("G21");
                        self.kirim("G91")
                        self.kirim("G92 Z0");
                        self.kirim("G1 Z0 F300")  # Kunci titik ini sebagai Titik Nol Z
                        time.sleep(0.5)
                        self.sensor_reader.bersihkan_event()
                        time.sleep(0.3)
                        break

                if not self.is_running: return
                if not pcb_tersentuh:
                    self.gagal("Probe belum menyentuh PCB; kalibrasi tidak dapat diselesaikan.")
                    return

                # 7. Kembalikan mesin ke pengaturan normal tanpa menunggu START fisik.
                self.status_update.emit("Kalibrasi selesai. Mulai program dari Jendela Utama.")

                # Kembalikan mesin ke pengaturan normal (Koordinat Absolut)
                self.kirim("$X");
                self.kirim("G21");
                self.kirim("G90");
                self.kirim("G17");
                self.kirim("G94")
                self.kirim("G92 X0 Y0 Z0");
                self.wait_grbl_ok()

                # Kirim sinyal sukses dan oper kabel USB-nya ke layar utama
                self.stop_sensor_reader()
                self.finished_success.emit(self.grbl, self.sensor)

            except Exception as e:
                # Jika di tengah jalan kabel kecabut
                print(f"Kalibrasi Hardware Terputus: {e}")
                self.gagal(f"Kalibrasi terputus: {e}")
        else:
            # JIKA TIDAK ADA MESIN (MODE SIMULASI UNTUK TESTING)
            if not self.is_running: return
            self.status_update.emit("[SIMULASI] Kalibrasi berjalan...")
            self.progress_update.emit(50)
            time.sleep(2.0)  # Pura-pura memakan waktu 2 detik
            self.status_update.emit("[SIMULASI] Selesai. Klik Jendela Utama.")
            self.progress_update.emit(100)
            self.finished_mock.emit()

    # Fungsi untuk menghentikan paksa pekerja ini
    def stop(self):
        self.is_running = False
        self.stop_sensor_reader()


# ==========================================
# JENDELA: LOGIN PENGGUNA (POP-UP PERTAMA)
# Tampilan untuk meminta password Admin123
# ==========================================
class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Panel Verifikasi Pengguna")
        self.setFixedSize(500, 350)
        self.setModal(True)  # Jendela tidak bisa di-klik di luarnya sebelum selesai
        self.setStyleSheet("background-color: #cccccc;")

        # Mengatur tumpukan tampilan dari atas ke bawah
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Hiasan Teks Atas (Kop Surat)
        header_frame = QFrame()
        header_frame.setStyleSheet("background-color: #d9d9d9; border-bottom: 2px solid black;")
        h_layout = QVBoxLayout(header_frame)
        lbl_h = QLabel("🪶 Jendela GUI Sistem Pengawasan dan Keamanan Mesin CNC 3 Axis untuk Milling PCB")
        lbl_h.setFont(QFont("Arial", 10))
        h_layout.addWidget(lbl_h)
        layout.addWidget(header_frame)

        # Judul Kotak Login
        header2_frame = QFrame()
        header2_frame.setStyleSheet("background-color: #cccccc; border-bottom: 2px solid black;")
        h2_layout = QVBoxLayout(header2_frame)
        lbl_panel = QLabel("PANEL VERIFIKASI PENGGUNA")
        lbl_panel.setFont(QFont("Arial", 14))
        lbl_panel.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h2_layout.addWidget(lbl_panel)
        layout.addWidget(header2_frame)

        # Area Formulir Isian
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(50, 20, 50, 20)
        content_layout.setSpacing(10)

        # Isian Nama (Boleh bebas)
        lbl_nama = QLabel("Masukkan Nama Dulu Ya")
        lbl_nama.setFont(QFont("Arial", 12))
        lbl_nama.setAlignment(Qt.AlignmentFlag.AlignCenter)
        content_layout.addWidget(lbl_nama)

        self.input_nama = QLineEdit()
        self.input_nama.setPlaceholderText("Isikan Nama")
        self.input_nama.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.input_nama.setStyleSheet("background-color: #a0a0a0; font-size: 12px; padding: 5px; border: none;")
        content_layout.addWidget(self.input_nama)

        content_layout.addSpacing(10)

        # Isian Password (Wajib "Admin123")
        lbl_pass = QLabel("<i>Password</i>-nya")
        lbl_pass.setFont(QFont("Arial", 12))
        lbl_pass.setAlignment(Qt.AlignmentFlag.AlignCenter)
        content_layout.addWidget(lbl_pass)

        self.input_pass = QLineEdit()
        self.input_pass.setPlaceholderText("Isikan Password")
        self.input_pass.setEchoMode(QLineEdit.EchoMode.Password)  # Teks berubah jadi bintang-bintang rahasia (*)
        self.input_pass.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.input_pass.setStyleSheet("background-color: #a0a0a0; font-size: 12px; padding: 5px; border: none;")
        content_layout.addWidget(self.input_pass)

        content_layout.addSpacing(20)

        # Tombol Kirim
        lbl_tanya = QLabel("Sudah?")
        lbl_tanya.setFont(QFont("Arial", 12))
        lbl_tanya.setAlignment(Qt.AlignmentFlag.AlignCenter)
        content_layout.addWidget(lbl_tanya)

        btn_ok = QPushButton("Klik Ok")
        btn_ok.setStyleSheet("""
            QPushButton { background-color: #e6e6e6; border: none; padding: 8px; font-size: 12px; }
            QPushButton:hover { background-color: #d9d9d9; }
        """)
        btn_ok.clicked.connect(self.cek_login)
        content_layout.addWidget(btn_ok)

        layout.addLayout(content_layout)
        layout.addStretch()

    # Mengecek apakah password sudah benar
    def cek_login(self):
        password = self.input_pass.text()
        if password == "Admin123":
            self.accept()  # Tutup jendela ini, ijinkan masuk
        else:
            QMessageBox.warning(self, "Login Gagal", "Password salah! Silakan coba lagi.")


# ==========================================
# JENDELA: PANDUAN KALIBRASI (POP-UP KEDUA)
# ==========================================
class CalibrationDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Proses Kalibrasi Mesin CNC")
        self.setFixedSize(650, 450)
        self.setModal(True)
        self.setStyleSheet("background-color: #cccccc;")

        self.grbl_serial = None
        self.sensor_serial = None
        self.is_mock = False
        self.calibration_active = False

        # Susunan tampilan (Sama seperti halaman login)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        header_frame = QFrame()
        header_frame.setStyleSheet("background-color: #d9d9d9; border-bottom: 2px solid black;")
        h_layout = QVBoxLayout(header_frame)
        lbl_h = QLabel("🪶 Jendela GUI Sistem Pengawasan dan Keamanan Mesin CNC 3 Axis untuk Milling PCB")
        lbl_h.setFont(QFont("Arial", 10))
        h_layout.addWidget(lbl_h)
        layout.addWidget(header_frame)

        header2_frame = QFrame()
        header2_frame.setStyleSheet("background-color: #cccccc; border-bottom: 2px solid black;")
        h2_layout = QVBoxLayout(header2_frame)
        lbl_panel = QLabel("PROSES KALIBRASI")
        lbl_panel.setFont(QFont("Arial", 14))
        lbl_panel.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h2_layout.addWidget(lbl_panel)
        layout.addWidget(header2_frame)

        # Teks Panduan Langkah-langkah
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(40, 20, 40, 20)

        # Kode HTML untuk membuat list nomor yang rapi
        instruksi = """
        <p style='font-size:14pt; margin-bottom:10px; text-align:center;'>Langkah-langkah melakukan kalibrasi :</p>
        <ol style='font-size:12pt; line-height:1.5;'>
            <li>Pasang PCB ke bracket</li>
            <li>Jepit probe merah ke bracket</li>
            <li>Jepit probe hitam ke mata pahat</li>
            <li>Klik tombol "KAL" dan tunggu hingga kalibrasi selesai dilakukan</li>
            <li>Proses kalibrasi tidak perlu menutup pintu body/casing</li>
            <li>Lepas probe hitam setelah kalibrasi selesai dilakukan. Jangan melepas probe merah yang telah terpasang!</li>
            <li>Tekan tombol "Jendela Utama" untuk masuk ke panel utama</li>
        </ol>
        """
        lbl_instruksi = QLabel(instruksi)
        lbl_instruksi.setWordWrap(True)  # Agar teks panjang turun ke bawah otomatis
        content_layout.addWidget(lbl_instruksi)

        # Teks status kecil di tengah
        self.lbl_status = QLabel("Menunggu aksi pengguna...")
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_status.setStyleSheet("color: blue; font-style: italic;")
        content_layout.addWidget(self.lbl_status)

        content_layout.addSpacing(20)

        # Tombol Navigasi Bawah
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(50)

        self.btn_kal = QPushButton("KAL")
        self.btn_kal.setFixedSize(200, 40)
        self.btn_kal.setStyleSheet("background-color: #e6e6e6; border: none; font-weight: bold; font-size: 12px;")
        self.btn_kal.clicked.connect(self.mulai_kalibrasi)

        self.btn_main = QPushButton("Jendela Utama")
        self.btn_main.setFixedSize(200, 40)
        self.btn_main.setEnabled(False)  # Dimatikan (redup) sampai mesin beres kalibrasi
        self.btn_main.setStyleSheet("background-color: #a0a0a0; border: none; font-size: 12px;")
        self.btn_main.clicked.connect(self.accept)

        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_kal)
        btn_layout.addWidget(self.btn_main)
        btn_layout.addStretch()

        content_layout.addLayout(btn_layout)
        layout.addLayout(content_layout)
        layout.addStretch()

        # Siapkan pekerja latar belakang (Tapi belum disuruh lari)
        self.worker = CalibrationWorker()
        self.worker.status_update.connect(self.lbl_status.setText)
        self.worker.finished_success.connect(self.on_success)
        self.worker.finished_mock.connect(self.on_mock)
        self.worker.failed.connect(self.on_failure)
        self.worker.finished.connect(self.on_worker_finished)

    def mulai_kalibrasi(self):
        # Kalau KAL diklik, matikan tombolnya biar user tidak klik dobel, lalu suruh pekerja lari
        self.calibration_active = True
        self.btn_kal.setEnabled(False)
        self.btn_kal.setStyleSheet("background-color: #a0a0a0; border: none;")
        self.worker.start()

    def on_success(self, grbl, sensor):
        # Tangkap dan simpan kabel USB dari pekerja latar belakang
        self.grbl_serial = grbl
        self.sensor_serial = sensor
        self.is_mock = False

        # Pekerjaan selesai! Hidupkan tombol Jendela Utama
        self.btn_main.setEnabled(True)
        self.btn_main.setStyleSheet("background-color: #e6e6e6; border: none; font-weight: bold;")

    def on_mock(self):
        # Mode Simulasi selesai
        self.is_mock = True
        self.btn_main.setEnabled(True)
        self.btn_main.setStyleSheet("background-color: #e6e6e6; border: none; font-weight: bold;")

    def on_failure(self, pesan):
        self.lbl_status.setText(pesan)
        self.btn_kal.setEnabled(True)
        self.btn_kal.setStyleSheet("background-color: #e6e6e6; border: none; font-weight: bold; font-size: 12px;")

    def on_worker_finished(self):
        self.calibration_active = False

    def closeEvent(self, event):
        if self.calibration_active or self.worker.isRunning():
            self.lbl_status.setText("Kalibrasi sedang berjalan. Jendela tidak dapat ditutup.")
            event.ignore()
            return
        event.accept()


# ==========================================
# JENDELA BARU: PLOT WINDOW (GAMBAR GRAFIK PCB)
# ==========================================
class PlotWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Visual Jalur PCB Real-Time")
        self.resize(550, 500)
        self.setStyleSheet("background-color: #cccccc;")
        layout = QVBoxLayout(self)

        # Membuat kanvas (papan tulis kosong) menggunakan Matplotlib
        self.fig = Figure(figsize=(5, 5), facecolor='#cccccc')
        self.canvas = FigureCanvas(self.fig)
        self.ax = self.fig.add_subplot(111)

        self.x_data, self.y_data = [], []
        self.setup_plot()

        layout.addWidget(self.canvas)

    # Fungsi untuk menyetel ulang papan tulis (menghapus coretan lama)
    def setup_plot(self):
        self.ax.clear()
        self.ax.set_title("Visual Jalur PCB", fontsize=12, fontweight='bold')
        self.ax.grid(True, linestyle='--', alpha=0.6)  # Tampilkan kotak-kotak pembantu
        self.ax.set_facecolor('#e6e6e6')

        # self.line = garis biru jalur (riwayat), self.point = titik merah (posisi saat ini)
        self.line, = self.ax.plot([], [], 'b-', linewidth=2)
        self.point, = self.ax.plot([], [], 'ro')
        self.x_data.clear()
        self.y_data.clear()
        self.canvas.draw()

    # Fungsi yang dipanggil berkali-kali setiap detik untuk menggambar titik baru
    def update_plot_data(self, tx, ty):
        # Agar tidak kebanyakan menggambar, gambar cuma kalau mesin pindah posisi lumayan jauh (0.02mm)
        if len(self.x_data) == 0 or (abs(tx - self.x_data[-1]) > 0.02 or abs(ty - self.y_data[-1]) > 0.02):
            self.x_data.append(tx)
            self.y_data.append(ty)

            # Sambungkan titik lama ke titik baru
            self.line.set_data(self.x_data, self.y_data)
            self.point.set_data([tx], [ty])

            # Auto-zoom layar agar seluruh gambar tetap terlihat penuh di bingkai
            self.ax.relim()
            self.ax.autoscale_view()
            self.canvas.draw()


# ==========================================
# JENDELA BARU: EMERGENCY POPUP (TANDA BAHAYA)
# ==========================================
class EmergencyPopup(QDialog):
    # Jendela ini menerima pesan alasan bahayanya (contoh: Pintu Terbuka!)
    def __init__(self, pesan_error="Mesin terdeteksi arus berlebih/kabel putus.", parent=None):
        super().__init__(parent)
        self.setWindowTitle("PERHATIAN!!")
        self.setFixedSize(450, 350)
        self.setModal(True)
        # Sembunyikan tanda (X) di pojok kanan atas agar user mau tak mau harus klik 'OK'
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.WindowTitleHint | Qt.WindowType.CustomizeWindowHint)
        self.setStyleSheet("background-color: #cccccc; border: 2px solid black;")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 10, 20, 20)
        layout.setSpacing(10)

        # Pasang ikon bahaya tegangan tinggi
        self.lbl_logo = QLabel()
        self.lbl_logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_logo.setStyleSheet("border: none;")
        pix_logo = QPixmap("logo listrik.png").scaled(64, 64, Qt.AspectRatioMode.KeepAspectRatio,
                                                      Qt.TransformationMode.SmoothTransformation)
        self.lbl_logo.setPixmap(pix_logo)
        layout.addWidget(self.lbl_logo)

        lbl_title = QLabel("EMERGENCY!!")
        lbl_title.setFont(QFont("Times New Roman", 20, QFont.Weight.Bold))
        lbl_title.setStyleSheet("color: red; border: none;")
        lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_title)

        lbl_subtitle = QLabel("SEGERA MATIKAN MESIN!!")
        lbl_subtitle.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        lbl_subtitle.setStyleSheet("color: black; border: none;")
        lbl_subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_subtitle)

        # Area Laporan Error (Kotak putih)
        frame_text = QFrame()
        frame_text.setStyleSheet("background-color: white; border: 1px solid black; border-radius: 2px;")
        frame_layout = QVBoxLayout(frame_text)

        text_report = f"<b>{pesan_error}</b><br><br>Klik OK untuk melakukan<br><b>KALIBRASI ULANG.</b>"
        self.lbl_report = QLabel(text_report)
        self.lbl_report.setFont(QFont("Arial", 10))
        self.lbl_report.setStyleSheet("color: black; border: none;")
        self.lbl_report.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_report.setWordWrap(True)
        frame_layout.addWidget(self.lbl_report)

        layout.addWidget(frame_text)
        layout.addSpacing(10)

        # Tombol OK Hijau Tua
        self.btn_ok = QPushButton("OK")
        self.btn_ok.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.btn_ok.setStyleSheet("""
            QPushButton { background-color: #2d862d; color: white; border: 2px solid black; padding: 10px; min-width: 100px; }
            QPushButton:hover { background-color: #246b24; }
            QPushButton:pressed { background-color: #1a4d1a; }
        """)
        self.btn_ok.clicked.connect(self.accept)
        layout.addWidget(self.btn_ok, alignment=Qt.AlignmentFlag.AlignCenter)


# ==========================================
# 1. PEKERJA LATAR BELAKANG: EKSEKUSI G-CODE
# Bertugas mengirim barisan teks GCode perlahan ke mesin
# ==========================================
class GRBLWorker(QThread):
    progress_update = pyqtSignal(int)
    status_update = pyqtSignal(str)
    sim_position_update = pyqtSignal(float, float)  # Untuk mengirim titik pura-pura saat simulasi

    def __init__(self, grbl_serial, sensor_serial, mock_mode):
        super().__init__()
        self.gcode_lines = []
        self.is_running = False
        self.grbl_serial = grbl_serial
        self.sensor_serial = sensor_serial
        self.mock_mode = mock_mode
        self.cur_x = 0.0
        self.cur_y = 0.0
        self.z_ketika_turun = False

    # Membaca file ".txt" lalu memasukkan seluruh isinya ke memori list
    def load_file(self, filepath):
        with open(filepath, 'r') as f:
            # Mengabaikan baris yang kosong atau hanya berupa komentar/catatan (berawalan kurung buka)
            lines = [l.strip() for l in f.readlines() if l.strip() and not l.startswith('(')]

        # Sesuai permintaan Anda dulu: Memastikan mesin langsung ke titik awal(nol) sebelum membaca file
        homing_sequence = ["G21", "G90", "G00 Z3.000", "G00 X0 Y0", "G00 Z0.000"]
        self.gcode_lines = homing_sequence + lines

    def wait_until_idle(self, timeout_seconds=120):
        """Tunggu sampai GRBL benar-benar berhenti bergerak sebelum menyatakan milling selesai."""
        deadline = time.monotonic() + timeout_seconds
        while self.is_running and time.monotonic() < deadline:
            try:
                self.grbl_serial.write(b"?\n")
                response = self.grbl_serial.readline().decode('utf-8', errors='ignore').strip()
            except Exception as error:
                print(f"Gagal membaca status akhir GRBL: {error}")
                return False
            if response:
                print(f"<< {response}")
            if response.startswith("<Idle"):
                return True
            if "error" in response.lower() or "alarm" in response.lower():
                return False
            time.sleep(0.1)
        return False

    def run(self):
        self.is_running = True
        self.status_update.emit('<i>On Progress</i>')
        total_lines = len(self.gcode_lines) if self.gcode_lines else 100

        # Membaca per baris dari awal sampai habis
        for i, line in enumerate(self.gcode_lines):
            if not self.is_running:
                break
            print(f">> {line}")  # Munculkan di terminal bawah

            # ----------------------------------------------------------------------------------
            # LOGIKA PENTING: MENGATUR SENSOR PUTARAN (ENCODER)
            # Jika barisan tulisan itu mengandung huruf Z, cek apakah angkanya minus (menembus PCB)
            # Jika minus, suruh Arduino Sensor untuk mulai membaca putaran motornya dengan kode #Z_TURUN;
            # ----------------------------------------------------------------------------------
            if not self.mock_mode and "Z" in line.upper():
                try:
                    for p in line.upper().split():
                        if p.startswith("Z"):
                            nilai_z = float(p[1:])
                            if nilai_z < 0 and not self.z_ketika_turun:
                                self.sensor_serial.write(b"#Z_TURUN;\n")
                                print("KIRIM #Z_TURUN;")
                                self.z_ketika_turun = True  # Kunci agar tidak dikirim berulang-ulang
                            elif nilai_z >= 0 and self.z_ketika_turun:
                                self.sensor_serial.write(b"#Z_NAIK;\n")  # Berhenti membaca putaran karena pahat di atas
                                print("KIRIM #Z_NAIK;")
                                self.z_ketika_turun = False
                except:
                    pass

            # Jika cuma simulasi, kita ektraksi paksa nilai X dan Y dari dalam teks untuk digambar di layar
            if self.mock_mode:
                match_x = re.search(r'[Xx]\s*([-+]?\d*\.?\d+)', line)
                match_y = re.search(r'[Yy]\s*([-+]?\d*\.?\d+)', line)
                if match_x: self.cur_x = float(match_x.group(1))
                if match_y: self.cur_y = float(match_y.group(1))
                if match_x or match_y:
                    self.sim_position_update.emit(self.cur_x, self.cur_y)
                time.sleep(0.05)

            # Jika asli, kita tembakkan baris itu ke kabel USB mesin
            if not self.mock_mode:
                cmd = (line + '\n').encode()  # Harus berbentuk kode biner (bytes)
                try:
                    self.grbl_serial.write(cmd)
                    # Mesin ini butuh sistem Ping-Pong, jangan lempar kode baru sebelum ada balasan "Ok"
                    while True:
                        res = self.grbl_serial.readline().decode('utf-8', errors='ignore').strip()
                        if res: print(f"<< {res}")
                        if res.lower() == "ok":
                            break
                        elif "error" in res.lower():
                            break
                except Exception as e:
                    print(f"Error Serial GRBL: {e}")

            # Hitung progress loading bar berdasarkan jumlah baris yang sudah dikirim
            persentase = int(((i + 1) / total_lines) * 100)
            self.progress_update.emit(persentase)

        # Kalau prosesnya selesai normal sampai akhir file
        if self.is_running and (self.mock_mode or self.wait_until_idle()):
            if not self.mock_mode:
                try:
                    self.sensor_serial.write(b"#MILLING_SELESAI;\n")  # Minta Arduino Sensor bunyikan lagu beres!
                except:
                    pass
            self.status_update.emit("<i>Done</i>")
        elif self.is_running:
            self.status_update.emit("<i>GRBL belum Idle; proses belum dinyatakan selesai.</i>")

    def stop(self):
        self.is_running = False


# ==========================================
# 2. PEKERJA LATAR BELAKANG: MEMANTAU SENSOR
# Ini bertugas menguping kabel USB yang terhubung ke Arduino Sensor (AR3)
# ==========================================
class SensorWorker(QThread):
    encoder_update = pyqtSignal(float, float)  # Kirim titik gambar nyata ke layar
    emergency_trigger = pyqtSignal(str)  # Teriakkan ada keadaan darurat ke layar
    feed_override_update = pyqtSignal(str)  # Beritahu layar bahwa tombol kecepatan diputar user

    def __init__(self, sensor_serial, mock_mode):
        super().__init__()
        self.is_running = True
        self.sensor_serial = sensor_serial
        self.mock_mode = mock_mode

    def run(self):
        while self.is_running:
            if not self.mock_mode:
                # Jika ada data yang datang dari USB Arduino
                if self.sensor_serial.in_waiting:
                    raw = self.sensor_serial.readline().decode('utf-8', errors='ignore').strip()
                    # Pastikan pesannya diawali # dan diakhiri ;
                    if raw.startswith("#") and raw.endswith(";"):
                        data = raw[1:-1]  # Kupas pagar dan titik-koma nya

                        # Jika ia ngirim posisi roda (contoh: POS,12,34)
                        if data.startswith("POS,"):
                            try:
                                _, tx, ty = data.split(',')
                                self.encoder_update.emit(float(tx), float(ty))  # Lempar angkanya ke tukang gambar
                            except:
                                pass

                        # Jika ia kasih kabar saklar kecepatan diputar
                        elif data in ["F_LOW", "F_MED", "F_HIGH"]:
                            self.feed_override_update.emit(data)

                        # Jika ia ngasih tanda bahaya mesin!
                        elif data == "LSBOFF":
                            self.emergency_trigger.emit("Bracket PCB Terlepas!")
                        elif data == "LSC_OFF":
                            self.emergency_trigger.emit("Casing Akrilik Terbuka!")
                        elif data == "STOP_ON":
                            self.emergency_trigger.emit("Tombol STOP Ditekan Secara Fisik!")

                    # Harus segera membalas "ACK" (Acknowledge) agar Arduino tidak cemberut nungguin
                    self.sensor_serial.write(b"ACK\n")
            else:
                time.sleep(0.5)

    def stop(self):
        self.is_running = False


# ==========================================
# 3. ANTARMUKA (GUI) PANEL PENGENDALI UTAMA
# Menyusun tata letak jendela pengawasan
# ==========================================
class CNCApp(QMainWindow):
    # Sinyal ini kita pakai buat minta komputer mereset ulang masuk ke menu Kalibrasi lagi
    request_recalibration = pyqtSignal()

    def __init__(self, grbl_serial, sensor_serial, is_mock):
        super().__init__()
        self.setWindowTitle("Sistem Pengawasan dan Keamanan Mesin CNC 3 Axis")
        self.setGeometry(100, 100, 650, 600)
        self.setStyleSheet("background-color: #cccccc;")

        # Status Penanda
        self.need_calibration = False  # Menandakan apakah user ngutang kalibrasi atau tidak
        self.waktu_mulai = None  # Pencatat jam main
        self.waktu_selesai = None  # Pencatat jam kelar
        self.grbl_serial_ref = grbl_serial  # Salinan kabel USB untuk menyuntik GCode kecepatan

        # Bikin Jendela Papan Tulis (tapi jangan ditampilin dulu)
        self.plot_window = PlotWindow()

        # Susun semua tombol & lampu
        self.setup_ui()
        # Nyalakan pekerja layar belakang
        self.start_threads(grbl_serial, sensor_serial, is_mock)

    def start_threads(self, grbl_serial, sensor_serial, is_mock):
        self.grbl_thread = GRBLWorker(grbl_serial, sensor_serial, is_mock)
        self.sensor_thread = SensorWorker(sensor_serial, is_mock)

        # Menyambungkan kabel-kabel komunikasi antara Pekerja Belakang ke Layar Depan
        self.grbl_thread.progress_update.connect(self.update_progress)
        self.grbl_thread.status_update.connect(self.update_status)

        # Kalau simulasi, yang gambar jalur ya GRBLWorker. Kalau asli, dari SensorWorker
        if is_mock:
            self.grbl_thread.sim_position_update.connect(self.plot_window.update_plot_data)
        else:
            self.sensor_thread.encoder_update.connect(self.plot_window.update_plot_data)

        # Sambungkan sinyal gawat darurat & warna kipas
        self.sensor_thread.emergency_trigger.connect(self.trigger_emergency)
        self.sensor_thread.feed_override_update.connect(self.update_spindle_gui)
        self.sensor_thread.start()

    # Bikin garis pembatas warna hitam horizontal
    def create_line(self):
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("border: 1px solid black;")
        return line

    # PROSES MENYUSUN TATA LETAK JENDELA UTAMA
    def setup_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        # Sistem Penataan Meja: Menggunakan "Grid" (baris dan kolom Excel-like) agar rapi
        grid = QGridLayout(main_widget)
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setSpacing(0)

        # 1. HEADER (Kop Surat paling atas) -> Letak: Baris 0, Panjang Kolom: 2
        header_frame = QFrame()
        header_frame.setStyleSheet("background-color: #d9d9d9; border-bottom: 2px solid black;")
        h_layout = QVBoxLayout(header_frame)
        lbl_h = QLabel("🪶 Jendela GUI Sistem Pengawasan dan Keamanan Mesin CNC 3 Axis untuk Milling PCB")
        lbl_h.setFont(QFont("Arial", 10))
        h_layout.addWidget(lbl_h)
        grid.addWidget(header_frame, 0, 0, 1, 2)

        # 2. HEADER PANEL -> Letak: Baris 1, Panjang Kolom: 2
        header2_frame = QFrame()
        header2_frame.setStyleSheet("background-color: #cccccc; border-bottom: 2px solid black;")
        h2_layout = QVBoxLayout(header2_frame)
        lbl_panel = QLabel("PANEL PENGAWASAN DAN KEAMANAN")
        lbl_panel.setFont(QFont("Arial", 14, QFont.Weight.Normal))
        lbl_panel.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h2_layout.addWidget(lbl_panel)
        grid.addWidget(header2_frame, 1, 0, 1, 2)

        # 3. PANEL KIRI (Deretan Lampu) -> Letak: Baris 2, Kolom 0
        left_panel = QFrame()
        left_panel.setFixedWidth(200)  # Terkunci ukuran 200 pixel lebarnya
        left_panel.setStyleSheet("background-color: #a6a6a6; border-right: 2px solid black;")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 20, 10, 20)

        lbl_indikator = QLabel("INDIKATOR")
        lbl_indikator.setFont(QFont("Arial", 12))
        lbl_indikator.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_indikator.setStyleSheet("border: none;")
        left_layout.addWidget(lbl_indikator)

        # Muat gambar-gambar dari memori penyimpanan
        self.pix_hijau_off = QPixmap("OFF.png").scaled(80, 80, Qt.AspectRatioMode.KeepAspectRatio,
                                                       Qt.TransformationMode.SmoothTransformation)
        self.pix_hijau_on = QPixmap("ON.png").scaled(80, 80, Qt.AspectRatioMode.KeepAspectRatio,
                                                     Qt.TransformationMode.SmoothTransformation)
        self.pix_merah_off = QPixmap("OFF-M.png").scaled(80, 80, Qt.AspectRatioMode.KeepAspectRatio,
                                                         Qt.TransformationMode.SmoothTransformation)
        self.pix_merah_on = QPixmap("ON-M.png").scaled(80, 80, Qt.AspectRatioMode.KeepAspectRatio,
                                                       Qt.TransformationMode.SmoothTransformation)
        self.pix_fan = QPixmap("balingbaling.png").scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                                          Qt.TransformationMode.SmoothTransformation)

        # Template membuat 1 Lampu + Teks Keterangan dibawahnya
        def add_lamp(layout, pix, text):
            lbl = QLabel();
            lbl.setPixmap(pix);
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter);
            lbl.setStyleSheet("border: none;")
            layout.addWidget(lbl)
            txt = QLabel(text);
            txt.setAlignment(Qt.AlignmentFlag.AlignCenter);
            txt.setFont(QFont("Arial", 9));
            txt.setStyleSheet("border: none;")
            layout.addWidget(txt)
            layout.addSpacing(10)
            return lbl

        self.lamp_start = add_lamp(left_layout, self.pix_hijau_off, "Mulai \"START\"")
        self.lamp_stop = add_lamp(left_layout, self.pix_merah_off, "BERHENTI \"STOP\"")
        self.lamp_emergency = add_lamp(left_layout, self.pix_merah_off, "EMERGENCY")

        # TOMBOL SIMULASI DARURAT YANG TRANSPARAN (DI TIMPA KE LAMPU DARURAT)
        self.btn_emergency_sim = QPushButton(self.lamp_emergency)
        self.btn_emergency_sim.setCursor(Qt.CursorShape.PointingHandCursor)  # Mouse jadi ikon telunjuk kalau kena
        self.btn_emergency_sim.setStyleSheet("background-color: transparent; border: none;")  # Sengaja tak kasat mata
        self.btn_emergency_sim.resize(80, 80)
        self.btn_emergency_sim.clicked.connect(self.simulate_emergency_click)

        left_layout.addStretch()
        grid.addWidget(left_panel, 2, 0)

        # 4. PANEL KANAN UTAMA -> Letak: Baris 2, Kolom 1
        right_panel = QWidget()
        right_panel.setStyleSheet("background-color: #cccccc;")
        r_layout = QVBoxLayout(right_panel)
        r_layout.setContentsMargins(20, 15, 20, 20)

        lbl_prog = QLabel("<i>PROGRESS BAR</i>")
        lbl_prog.setFont(QFont("Arial", 12))
        lbl_prog.setAlignment(Qt.AlignmentFlag.AlignCenter)
        r_layout.addWidget(lbl_prog)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(25)
        # Memaksa Progress Bar untuk elastis ke kanan mentok dinding
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("""
            QProgressBar { border: 1px solid gray; background-color: #b3b3b3; text-align: center; color: black; font-size: 12px; }
            QProgressBar::chunk { background-color: #1aa31a; margin: 0px; }
        """)
        r_layout.addWidget(self.progress_bar)
        r_layout.addSpacing(20)

        self.lbl_status = QLabel("STATUS MESIN : <i>Off Condition</i>")
        self.lbl_status.setFont(QFont("Arial", 16))
        r_layout.addWidget(self.lbl_status)
        r_layout.addSpacing(20)

        # Area Tombol Cerdas (Rata Kanan)
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()  # Mendorong semua tombol ke sebelah kanan dengan "per" transparan
        btn_vbox = QVBoxLayout()
        btn_style = """
            QPushButton { background-color: #a0a0a0; border: none; padding: 10px; font-size: 11px; min-width: 150px; }
            QPushButton:hover { background-color: #8c8c8c; }
            QPushButton:pressed { background-color: #737373; }
        """
        btn_load = QPushButton("MULAI PROGRAM")
        btn_load.setStyleSheet(btn_style)
        btn_load.clicked.connect(self.load_and_start)

        self.btn_excel = QPushButton("Simpan Data .xs")
        self.btn_excel.setStyleSheet(btn_style)
        self.btn_excel.clicked.connect(self.save_to_excel)

        btn_vbox.addWidget(btn_load)
        btn_vbox.addSpacing(10)
        btn_vbox.addWidget(self.btn_excel)
        btn_layout.addLayout(btn_vbox)
        r_layout.addLayout(btn_layout)

        r_layout.addSpacing(15)
        r_layout.addWidget(self.create_line())
        r_layout.addSpacing(15)

        # Area Indikator Kipas Spindle
        lbl_spindle = QLabel("STATUS KECEPATAN SPINDLE")
        lbl_spindle.setFont(QFont("Arial", 14))
        lbl_spindle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        r_layout.addWidget(lbl_spindle)

        spindle_hbox = QHBoxLayout()
        spindle_hbox.setAlignment(Qt.AlignmentFlag.AlignCenter)
        spindle_hbox.setSpacing(40)  # Jarak antar kipas 40px
        r_layout.addLayout(spindle_hbox)

        # Template membuat 1 Kipas + Kotak Warna di bawahnya yang saling menempel (Spacing: 0)
        def create_fan_widget(text):
            vbox = QVBoxLayout()
            vbox.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.setSpacing(0)
            vbox.setContentsMargins(0, 0, 0, 0)
            fan = QLabel()
            fan.setAlignment(Qt.AlignmentFlag.AlignCenter)
            fan.setPixmap(self.pix_fan)
            lbl = QLabel(text)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setFont(QFont("Arial", 10))
            lbl.setStyleSheet("background-color: grey; color: black; padding: 3px 15px;")  # Mati = Warna Abu
            vbox.addWidget(fan)
            vbox.addWidget(lbl)
            return vbox, lbl

        fan_rendah_layout, self.lbl_rendah = create_fan_widget("RENDAH")
        fan_sedang_layout, self.lbl_sedang = create_fan_widget("SEDANG")
        fan_tinggi_layout, self.lbl_tinggi = create_fan_widget("TINGGI")

        spindle_hbox.addLayout(fan_rendah_layout)
        spindle_hbox.addLayout(fan_sedang_layout)
        spindle_hbox.addLayout(fan_tinggi_layout)

        r_layout.addStretch()  # Dorong area abu ke bawah agar layar proporsional
        grid.addWidget(right_panel, 2, 1)  # Pasang panel kanan ini ke grid utama kolom 1

    # FUNGSI RESPONS WARNA KIPAS SPINDLE
    # Berdasarkan pesan rahasia yang ditangkap oleh SensorWorker
    def update_spindle_gui(self, mode):
        # Matikan semua lampunya jadi abu-abu dulu
        self.lbl_rendah.setStyleSheet("background-color: grey; color: black; padding: 3px 15px;")
        self.lbl_sedang.setStyleSheet("background-color: grey; color: black; padding: 3px 15px;")
        self.lbl_tinggi.setStyleSheet("background-color: grey; color: white; padding: 3px 15px;")

        try:
            if mode == "F_LOW":
                # Nyalakan indikator hijau terang
                self.lbl_rendah.setStyleSheet("background-color: #00FF00; color: black; padding: 3px 15px;")
                if not self.grbl_thread.mock_mode:
                    # Suntik kode Hex Override ke kabel USB mesin GRBL secara real-time
                    self.grbl_serial_ref.write(bytes([0x90]))  # Reset ke 100% bawaan
                    time.sleep(0.05)
                    for _ in range(5): self.grbl_serial_ref.write(bytes([0x92]))  # Tekan pelan 5 kali (Jadi 50%)
            elif mode == "F_MED":
                self.lbl_sedang.setStyleSheet("background-color: yellow; color: black; padding: 3px 15px;")
                if not self.grbl_thread.mock_mode:
                    self.grbl_serial_ref.write(bytes([0x90]))
                    time.sleep(0.05)
                    for _ in range(5): self.grbl_serial_ref.write(bytes([0x91]))  # Tekan cepat 5 kali (Jadi 150%)
            elif mode == "F_HIGH":
                self.lbl_tinggi.setStyleSheet("background-color: red; color: white; padding: 3px 15px;")
                if not self.grbl_thread.mock_mode:
                    self.grbl_serial_ref.write(bytes([0x90]))  # Cukup biarkan normal mesin melaju keras 100%
        except:
            pass

    # Jika Tombol "Jendela Jalur PCB" diklik
    def load_and_start(self):
        # Jika user ketahuan ngutang kalibrasi karena habis beres PCB sebelumnya, tolak dengan keras!
        if self.need_calibration:
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setWindowTitle("Kalibrasi Dibutuhkan")
            msg.setText(
                "Mesin telah menyelesaikan milling sebelumnya.\nAnda harus melakukan kalibrasi ulang sebelum memuat file G-Code baru.")
            btn_kal = msg.addButton("Kalibrasi Sekarang", QMessageBox.ButtonRole.ActionRole)
            btn_batal = msg.addButton("Batal", QMessageBox.ButtonRole.RejectRole)
            msg.exec()

            if msg.clickedButton() == btn_kal:
                self.request_recalibration.emit()  # Lempar sinyal buat restart ke Jendela Kalibrasi
            return

        # Program milling sudah ditetapkan agar operator tidak perlu memilih file lagi.
        filepath = DEFAULT_GCODE_FILE
        if filepath.is_file():
            # Re-Inisiasi pekerja kalau kebetulan dia lagi tidur
            if not self.grbl_thread.isRunning():
                old_mock = self.grbl_thread.mock_mode
                old_grbl = self.grbl_thread.grbl_serial
                old_sensor = self.sensor_thread.sensor_serial
                self.grbl_thread = GRBLWorker(old_grbl, old_sensor, old_mock)
                self.grbl_thread.progress_update.connect(self.update_progress)
                self.grbl_thread.status_update.connect(self.update_status)

                if old_mock:
                    self.grbl_thread.sim_position_update.connect(self.plot_window.update_plot_data)

            # Suruh pekerja menelan isi file text tersebut ke memorinya
            self.grbl_thread.load_file(filepath)

            # Tampilkan layar gambar ke samping/layar sekunder
            self.plot_window.setup_plot()
            self.plot_window.show()

            # Bersihkan lampu & grafik jadi mode hijau segar (Siap Berangkat)
            self.lamp_start.setPixmap(self.pix_hijau_on)
            self.lamp_stop.setPixmap(self.pix_merah_off)
            self.lamp_emergency.setPixmap(self.pix_merah_off)
            self.progress_bar.setValue(0)

            # Mulai mencatat jam kerja pakai datetime
            self.waktu_mulai = datetime.datetime.now()
            self.waktu_selesai = None

            if not self.grbl_thread.mock_mode:
                try:
                    self.sensor_thread.sensor_serial.write(b"#SPINDLE_ON;\n")  # Minta bor dinyalakan
                except:
                    pass

            # PECUT pekerja GRBL, "AYO JALAN!"
            self.grbl_thread.start()
        else:
            QMessageBox.critical(self, "Program Tidak Ditemukan", f"File G-Code tidak ditemukan:\n{filepath}")

    # Fungsi menerima lapor status loading dari pekerja untuk dipasang ke grafik GUI
    def update_progress(self, val):
        self.progress_bar.setValue(val)

    def update_status(self, status):
        self.lbl_status.setText(f"STATUS MESIN : {status}")
        self.lbl_status.setStyleSheet("color: black;")

        # KALAU TUGAS SUDAH SAMPAI BAWAH (100% / DONE)
        if "Done" in status:
            self.waktu_selesai = datetime.datetime.now()  # Catat jam selesai
            self.lamp_start.setPixmap(self.pix_hijau_off)  # Matikan lampu jalan

            self.need_calibration = True  # Tandai ngutang kalibrasi!

            if not self.grbl_thread.mock_mode:
                try:
                    self.sensor_thread.sensor_serial.write(b"#SPINDLE_OFF;\n")
                except:
                    pass

            # Pilihan untuk user, mau nge-save Excel dulu (OK), atau langsung dihajar Reset (Kalibrasi)?
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Proses Selesai")
            msg.setText(
                "Proses milling telah selesai (Done).\n\nKlik 'Kalibrasi' jika Anda ingin langsung mereset mesin sekarang.\nKlik 'OK' jika Anda ingin menyimpan data (.xlsx) terlebih dahulu.")
            btn_kal = msg.addButton("Kalibrasi", QMessageBox.ButtonRole.ActionRole)
            btn_ok = msg.addButton("OK", QMessageBox.ButtonRole.ActionRole)
            msg.exec()

            if msg.clickedButton() == btn_kal:
                self.request_recalibration.emit()

    def simulate_emergency_click(self):
        self.trigger_emergency("Simulasi GUI Tombol Emergency")

    # FUNGSI YANG JALAN JIKA KABEL MESIN KETARIK / PINTU DIBUKA PAKSA / ARUS KELEBIHAN
    def trigger_emergency(self, alasan):
        self.grbl_thread.stop()  # Hentikan seketika pekerja yang sedang mengirim barisan Gcode
        self.waktu_selesai = datetime.datetime.now()

        # Ubah layar status jadi warna merah mencolok
        self.lbl_status.setText("STATUS MESIN : <i>Emergency</i>")
        self.lbl_status.setStyleSheet("color: red; font-weight: bold;")

        # Nyalakan lampu sirine mesin
        self.lamp_emergency.setPixmap(self.pix_merah_on)
        self.lamp_start.setPixmap(self.pix_hijau_off)
        self.lamp_stop.setPixmap(self.pix_merah_off)

        # Suntik kode rahasia \x18 (Soft Reset GRBL) ke kabel mesin
        # Efeknya semua step motor nge-rem dadakan tapi koordinat nggak lupa
        if not self.grbl_thread.mock_mode:
            try:
                self.sensor_thread.sensor_serial.write(b"#SPINDLE_OFF;\n")
                self.grbl_thread.grbl_serial.write(b"!")  # Jeda / Hold
                time.sleep(0.1)
                self.grbl_thread.grbl_serial.write(b"\x18")  # Reset
                time.sleep(0.5)
                self.grbl_thread.grbl_serial.write(b"$X\n")  # Unlock
            except:
                pass

        # Sembunyikan layar TV Grafik sementara
        self.plot_window.hide()
        self.grbl_thread.wait()

        # Munculkan Popup Peringatan Bergambar Bahaya
        popup = EmergencyPopup(pesan_error=alasan, parent=self)
        if popup.exec() == QDialog.DialogCode.Accepted:  # Jika OK terpaksa dipencet
            self.request_recalibration.emit()  # Program kembali restart ke Halaman Kalibrasi

    # Fitur Pembuat Laporan Bukti Kerja dalam format Microsoft Excel (.xlsx)
    def save_to_excel(self):
        # Mencegah user yang iseng nge-save tapi mesin belum jalan apa-apa
        if not self.waktu_mulai:
            QMessageBox.warning(self, "Peringatan",
                                "Mesin belum pernah dijalankan!\nSilakan lakukan proses milling terlebih dahulu.")
            return

        # MENCEGAH PENCURIAN DATA SEBELUM MATANG
        # Kalau jam mulai sudah ada, tapi mesin belum mencatat jam berhenti, artinya mesin masih asik ngukir! Tolak!
        if self.waktu_mulai and not self.waktu_selesai:
            QMessageBox.warning(self, "Peringatan",
                                "Mesin sedang beroperasi!\nHarap tunggu hingga proses milling selesai (Done) atau dihentikan untuk menyimpan data.")
            return

        # Buka dialog save as Windows
        filepath, _ = QFileDialog.getSaveFileName(self, "Simpan Laporan", "Laporan_Milling_CNC.xlsx",
                                                  "Excel Files (*.xlsx)")
        if not filepath: return

        try:
            waktu_akhir = self.waktu_selesai if self.waktu_selesai else datetime.datetime.now()

            # Buat file Excel Baru dari nol
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan CNC"

            # Tulis Teks ke dalam kotak-kotak sel excel
            ws['A1'] = "LAPORAN PENGAWASAN MESIN CNC 3 AXIS"
            ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            ws['A3'] = "Waktu Mulai:";
            ws['B3'] = self.waktu_mulai.strftime("%d %B %Y, %H:%M:%S")
            ws['A4'] = "Waktu Selesai:";
            ws['B4'] = waktu_akhir.strftime("%d %B %Y, %H:%M:%S")
            ws['A6'] = "Visualisasi Jalur PCB:"

            # Screenshot diam-diam grafik layar sekunder, lalu tempel di atas file excel (Sel A7)
            temp_img_path = "temp_plot_cnc.png"
            self.plot_window.fig.savefig(temp_img_path, dpi=150, bbox_inches='tight')

            img = ExcelImage(temp_img_path)
            ws.add_image(img, 'A7')

            # Simpan Berkas
            wb.save(filepath)

            # Buang jejak file fotonya biar memory nggak penuh
            if os.path.exists(temp_img_path): os.remove(temp_img_path)
            QMessageBox.information(self, "Berhasil", f"Laporan berhasil disimpan ke:\n{filepath}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan file Excel:\n{e}")

    # Fungsi kalau user nekat mengeklik tombol X silang merah di sudut kanan atas Windows
    def closeEvent(self, event):
        self.plot_window.close()
        self.grbl_thread.stop()
        self.sensor_thread.stop()

        if hasattr(self.grbl_thread, 'grbl_serial') and self.grbl_thread.grbl_serial:
            try:
                self.grbl_thread.grbl_serial.close()
            except:
                pass
        if hasattr(self.sensor_thread, 'sensor_serial') and self.sensor_thread.sensor_serial:
            try:
                self.sensor_thread.sensor_serial.close()
            except:
                pass
        event.accept()


# ==========================================
# BOOT SEQUENCE (LURUNG LOGIKA AWAL APLIKASI)
# Mengatur siapa duluan yang muncul ke layar!
# ==========================================
if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        continue_program = True

        # Kita pakai sistem "Putaran Kehidupan" (While Loop)
        while continue_program:

            # TAHAP 1: PINTU GERBANG (LOGIN)
            login_dialog = LoginDialog()
            # Munculkan ke Layar (Tertahan di sini sampai pass 'Admin123' terisi)
            if login_dialog.exec() == QDialog.DialogCode.Accepted:

                # TAHAP 2: JENDELA INSTRUKSI KALIBRASI
                cal_dialog = CalibrationDialog()
                # Jika user klik Jendela Utama...
                if cal_dialog.exec() == QDialog.DialogCode.Accepted:

                    # TAHAP 3: BUKA JENDELA KENDALI MESIN
                    window = CNCApp(cal_dialog.grbl_serial, cal_dialog.sensor_serial, cal_dialog.is_mock)

                    # Titip status bohong (False) dulu. Berguna nanti kalau emergency terjadi
                    restart_requested = [False]


                    # Fungsi buat disuruh nutup jendela
                    def handle_recalc_request():
                        restart_requested[0] = True
                        window.close()

                        # Pasang kabel sinyal reset ke tombol tadi


                    window.request_recalibration.connect(handle_recalc_request)

                    window.show()

                    # Nyalakan Nyawa Aplikasi Windows-nya (Aplikasi diem ngegantung di sini sampai dimatikan)
                    app.exec()

                    # --------- JIKA TIBA-TIBA APLIKASINYA MATI ---------

                    # Apakah ia mati karena sinyal kalibrasi ulang (Emergency/Reset)?
                    if restart_requested[0]:
                        if hasattr(window, 'plot_window'): window.plot_window.close()
                        del window
                        # Looping akan berputar, naik kembali ke Tahap 1, TAPI...
                        # Karena kita pakai Logika "While" pintar, ia justru melompat ke TAHAP 2 (Kalibrasi Langsung), tanpa minta Password Ulang!
                        continue

                        # Atau matinya emang karena user menekan (X) untuk pulang?
                    else:
                        continue_program = False
                        break

                else:  # Jika di halaman Kalibrasi malah diklik (X)
                    continue_program = False
                    break
            else:  # Jika di halaman Login malah diklik (X)
                print("[SISTEM] Verifikasi dibatalkan pengguna.")
                break

    except Exception as e:
        print(f"GAGAL MEMULAI APLIKASI: {e}")
